from multiprocessing.heap import Arena
from tkinter import*
from tkinter import messagebox
from PIL import Image, ImageTk
import cv2
import numpy as np
from playsound import playsound
import time
import xlsxwriter
import datetime
import openpyxl
from cvzone.SerialModule import SerialObject


img_counter = 0
count = 0
globalAREA = 0 
stop= 0 
wri = 1 
WeightStatus = 0 
width = 460
height = 680
active = 'disable'
#---------

workbook = xlsxwriter.Workbook('mega1\\Resur\\Area.xlsx')
worksheet = workbook.add_worksheet("Sheet1")

#---------
mainWindow = Tk()
mainWindow.geometry("1920x1080")
mainWindow.title('Cân Tôm  by VP')
icon = PhotoImage(file='mega1\\icon\\icon.png')
mainWindow.iconphoto(True,icon) 
mainWindow.configure(bg='black')
labels = Label(mainWindow,text= "CÂN TÔM XỬ LÝ ẢNH : NHÓM 5 ",font=('Arial',20,'bold'),fg='#96ed5c',bg='black',relief=SUNKEN,bd=4,padx=25,pady=8)
labels.place(x=500, y=5)
idcam = Label(mainWindow,text= "Camera ",font=('Arial',12,'bold'),fg='#96ed5c',bg='black')
idcam.place(x=50, y=682)
idcam = Label(mainWindow,text= "Enter ID Camera  ",font=('Arial',12,'bold'),fg='#96ed5c',bg='black')
idcam.place(x=265, y=662)



#-------------------- Frame  mainWindow -----------------
# Fram 1 
cam_on = False
cap = None
vdiFrame = Frame(mainWindow,bg= 'white', height =0, width =0 )
vdiFrame.place(x=5,y=65)
vid_lbl = Label(vdiFrame)
vid_lbl.grid(row=0, column=0)
vid_lbl1 = Label(vdiFrame)
vid_lbl1.grid(row=1, column=0)

vdiFrame1 = Frame(mainWindow,bg= 'white', height =0, width =0)
vdiFrame1.place(x=300,y=65)
vid_lbl2 = Label(vdiFrame1)
vid_lbl2.grid(row=0, column=0)
vid_lbl3 = Label(vdiFrame1)
vid_lbl3.grid(row=0, column=1)


#----------------Value Setting -------------------
#-------------------------
def show_frame():
    book = openpyxl.load_workbook('mega1\\Setting\\Setting.xlsx')
    sheet = book.active
    global  camer , cm 
    global S1,S2,S3,S4,S5,S6,S7,S8
    S1 = sheet['B2']
    S1 =S1.value
    S2 = sheet['C2']
    S2 =S2.value
    S3 = sheet['D2']
    S3 =S3.value
    S4 = sheet['B3']
    S4 =S4.value
    S5 = sheet['C3']
    S5 =S5.value
    S6 = sheet['D3']
    S6 =S6.value
    S7 = sheet['B4']
    S7 =S7.value
    S8 = sheet['C4']
    S8 =S8.value

    if cam_on:
        global res, imgDil 
        success, imgvid = cap.read()
        value()

        imgvid =imgvid[y1:y2,x1:x2]
        camer =cv2.cvtColor(imgvid, cv2.COLOR_BGR2RGB)
        cm = cv2.cvtColor(imgvid, cv2.COLOR_BGR2RGB)
        imgContour1= camer
        imgBlur = cv2.GaussianBlur(camer,(7,7),1)
        hsv = cv2.cvtColor(imgBlur, cv2.COLOR_BGR2HSV)
        hsv= cv2.GaussianBlur(hsv,(7,7),1)
        lower_blue = np.array([S1,S2,S3])
        upper_blue = np.array([S4,S5,S6])
        mask = cv2.inRange(hsv, lower_blue, upper_blue)
        res = cv2.bitwise_and(imgBlur,imgBlur, mask= mask)
        imgCanny = cv2.Canny(mask ,S7,S8)
        kernel = np.ones((5,5))
        imgDil = cv2.dilate(imgCanny,kernel,iterations=1 )
        line(imgDil,imgContour1)
        caps()
        imgv = Image.fromarray(res).resize((300,210))
        imgtk = ImageTk.PhotoImage(image=imgv)  
        vid_lbl.imgtk = imgtk    
        vid_lbl.configure(image=imgtk)    
   
        imgv = Image.fromarray(imgDil).resize((300,210))
        imgtk = ImageTk.PhotoImage(image=imgv)  
        vid_lbl1.imgtk = imgtk    
        vid_lbl1.configure(image=imgtk)  

        imgv = Image.fromarray(imgContour1).resize((640,480))
        imgtk = ImageTk.PhotoImage(image=imgv)  
        vid_lbl2.imgtk = imgtk    
        vid_lbl2.configure(image=imgtk)  

        mainWindow.after(20, show_frame)

def showResur():
    hinh = "mega1\\Resur\\Image\\id_{}_".format(img_counter)+"opencv_frame_{}.png".format(img_counter)
    imgs  = cv2.imread(hinh)
    h =  imgs.shape[1]
    w = imgs.shape[0]
    imgContour= imgs.copy()
    imgBlur = cv2.GaussianBlur(imgContour,(7,7),1)
    imgCanny = cv2.Canny(imgBlur,10,10)
    kernel = np.ones((5,5))
    imgDil1 = cv2.dilate(imgCanny,kernel,iterations=1 )
    getContours(imgDil1,imgContour) 
    imgv = Image.fromarray(imgContour).resize(((h,w)))
    imgtk = ImageTk.PhotoImage(image=imgv)  
    vid_lbl3.imgtk = imgtk    
    vid_lbl3.configure(image=imgtk)


#-- CONNET COM ----# 
def connet():
    global arduino
    books = openpyxl.load_workbook('mega1\\Setting\\Com.xlsx')
    ser = books.active
    coms = ser['A1']
    coms = coms.value
    print(coms) 
    arduino = SerialObject(coms)
connet()
def ons():
    arduino.sendData([1])
def ofs():
    arduino.sendData([4])
def couter():
    arduino.sendData([2])
def uncouter():
    arduino.sendData([3])

#-----------------------Start Stop Camera -----------------
def start_vid():
    global cam_on, cap
    #stop_vid()
    cam_on = True
    cap = cv2.VideoCapture(idcam) 
    show_frame()
    ons()
    
 
def stop_vid():
    global cam_on
    global stop 
    cam_on = False
    ofs()
    if cap:
        cap.release()
#--------------------------------------------------------




#################### BUTTON #######################
on = PhotoImage(file="mega1\\icon\\switch-on.png")
on = on.zoom(12)
on = on.subsample(52)
off = PhotoImage(file="mega1\\icon\\switch-off.png")
off = off.zoom(12)
off= off.subsample(52)
switch_value = True
def buttonCAMERA():
    def toggle():
        global switch_value
        if switch_value == True:
            button.config(image=on, bg="black",bd=0,
                        activebackground="black")
            
            start_vid()
        
            switch_value = False
    
        else:
            button.config(image=off, bg="black", bd=0,
                        activebackground="black")
            
            stop_vid()
        
            switch_value = True

    button = Button(mainWindow,image=off,bd=0 ,state=active, command=toggle,bg='black',activebackground='black')
    button.place(x=25, y=700)

######################## CHECK BUTTON #########################
def display():
    global active , idcam
    if(x.get()==1):
        print('you choose camera 0')
        active ='active'
        idcam = 0
        print('Camera 0 on ')
        buttonCAMERA()
    if(x.get()==2):
        print('you choose camera 1')
        active ='active'
        idcam = 1
        print('Camera 1 on ')
        buttonCAMERA()
    if(x.get()==4):
        print('you choose camera 2')
        active ='active'
        idcam = 2
        print('Camera 1 on ')
        buttonCAMERA()
    if(x.get()==3):
        print('No ID camera , plee chosse :(')
        active ='disable'
        stop_vid()
        buttonCAMERA()
x = IntVar()
ckeck_button1 =  Checkbutton(mainWindow,text='Camera 0 ',variable=x,onvalue=1,offvalue=3,command=display,font=('Arial',10),fg= '#3a32a8',bg= '#b8b8e3')
ckeck_button1.place(x=270, y=690)
ckeck_button2 =  Checkbutton(mainWindow,text='Camera 1 ',variable=x,onvalue=2,offvalue=3,command=display,font=('Arial',10),fg= '#3a32a8',bg='#b8b8e3')
ckeck_button2.place(x=370, y=690)
ckeck_button3 =  Checkbutton(mainWindow,text='Camera 2 ',variable=x,onvalue=4,offvalue=3,command=display,font=('Arial',10),fg= '#3a32a8',bg='#b8b8e3')
ckeck_button3.place(x=470, y=690)



#-----------------------------SETTING---------------------------------------------------------------------------------------------------
def setting():
    
    if active =='disable':
        if messagebox.showerror(title="Erron",message="Camera ID not selected"):
            print(idcam)
    
    else:
        if messagebox.showwarning(title="Warnig",message="The program may crash while performing this task"):
            book = openpyxl.load_workbook('mega1\\Setting\\Setting.xlsx')
            sheet = book.active
            workbook1 = xlsxwriter.Workbook('mega1\\Setting\\Setting.xlsx')
            worksheet1 = workbook1.add_worksheet("Setting")

            newSetting = Toplevel(mainWindow)
            frame = Frame(newSetting , bg= 'black',width=350,height=300)
            frame.place(x=0, y=0)
            F_lbl = Label(frame)
            F_lbl.grid(row=0, column=0)

            res = Frame(newSetting , bg= 'black',width=350,height=300)
            res.place(x=352, y=0)
            R_lbl = Label(res)
            R_lbl.grid(row=0, column=0)

            dil = Frame(newSetting , bg= 'black',width=350,height=300)
            dil.place(x=704, y=0)
            D_lbl = Label(dil)
            D_lbl.grid(row=0, column=0)
 
            #--------- COM --------------
            global com
            def textCOM():
                books = openpyxl.load_workbook('mega1\\Setting\\Com.xlsx')
                ser = books.active
                coms = ser['A1']
                coms = coms.value
                textCOM = Label(newSetting,text=coms)
                textCOM.place(x = 0 , y=520)
            textCOM()
            com= Entry(newSetting,font=("Arial",15))
            com.place(x=0,y=550)
            def wrs(e):
                global com
                workbook = xlsxwriter.Workbook('mega1\\Setting\\Com.xlsx')
                worksheet = workbook.add_worksheet("Sheet1")
                worksheet.write('A1' , str(com.get()))
                workbook.close()
                connet()
                textCOM()
                com.delete(0, 'end')

            newSetting.bind('<Return>',wrs)


            stop_vid()
            def show():
                if cam_on:
                    ret, imgvid = cap.read()
                    value()
                    imgvid =imgvid[y1:y2,x1:x2]
                    camer =cv2.cvtColor(imgvid, cv2.COLOR_BGR2RGB)
                    imgBlur = cv2.GaussianBlur(camer,(7,7),1)
                    hsv = cv2.cvtColor(imgBlur, cv2.COLOR_BGR2HSV)
                    hsv= cv2.GaussianBlur(hsv,(7,7),1)
                    lower_blue = np.array([scale1.get(),scale2.get(),scale3.get() ])
                    upper_blue = np.array([scale4.get(),scale5.get(),scale6.get() ])
                    mask = cv2.inRange(hsv, lower_blue, upper_blue)
                    res = cv2.bitwise_and(imgBlur,imgBlur, mask= mask)
                    imgCanny = cv2.Canny(res ,scale7.get(),scale8.get())
                    kernel = np.ones((5,5))
                    imgDil = cv2.dilate(imgCanny,kernel,iterations=1 )

                    imgv = Image.fromarray(camer).resize((348,298))
                    imgtk = ImageTk.PhotoImage(image=imgv) 
                    F_lbl.imgtk = imgtk    
                    F_lbl.configure(image=imgtk)  

                    #imgv = Image.fromarray(mark).resize((348,298))
                    #imgtk = ImageTk.PhotoImage(image=imgv) 
                    #M_lbl.imgtk = imgtk    
                    #M_lbl.configure(image=imgtk)  

                    imgv = Image.fromarray(res).resize((348,298))
                    imgtk = ImageTk.PhotoImage(image=imgv) 
                    R_lbl.imgtk = imgtk    
                    R_lbl.configure(image=imgtk) 

                    imgv = Image.fromarray(imgDil).resize((348,298))
                    imgtk = ImageTk.PhotoImage(image=imgv) 
                    D_lbl.imgtk = imgtk    
                    D_lbl.configure(image=imgtk)  
                    newSetting.after(20,show)
            def start():
                global cam_on, cap
                #stop_vid()
                cam_on = True
                cap = cv2.VideoCapture(idcam) 
                show()
            #--------button-------------
            start= Button(newSetting,text='Start Camera',command=start,width=12,height=2)
            start.place(x=960,y=550)
            #---------Scale-------------
            scale1 = Scale(newSetting,from_=0 , to=700,length=200,orient=HORIZONTAL,troughcolor='#cbcf17',fg='#cf1729')
            book = openpyxl.load_workbook('mega1\\Setting\\Setting.xlsx')
            sheet = book.active
            S1 = sheet['B2']
            scale1.set(S1.value)
            scale1.place(x=355,y=360)

            scale2 = Scale(newSetting,from_=0 , to=700,length=200,orient=HORIZONTAL,troughcolor='#cbcf17',fg='#cf1729')
            S2 = sheet['C2']
            scale2.set(S2.value)
            scale2.place(x=570,y=360)

            scale3 = Scale(newSetting,from_=0 , to=700,length=200,orient=HORIZONTAL,troughcolor='#cbcf17',fg='#cf1729')
            S3 = sheet['D2']
            scale3.set(S3.value)
            scale3.place(x=785,y=360)

            scale4 = Scale(newSetting,from_=0 , to=700,length=200,orient=HORIZONTAL,troughcolor='#cbcf17',fg='#cf1729')
            S4 = sheet['B3']
            scale4.set(S4.value)
            scale4.place(x=355,y=450)

            scale5 = Scale(newSetting,from_=0 , to=700,length=200,orient=HORIZONTAL,troughcolor='#cbcf17',fg='#cf1729')
            S5 = sheet['C3']
            scale5.set(S5.value)
            scale5.place(x=570,y=450)

            scale6= Scale(newSetting,from_=0 , to=700,length=200,orient=HORIZONTAL,troughcolor='#cbcf17',fg='#cf1729')
            S6 = sheet['D3']
            scale6.set(S6.value)
            scale6.place(x=785,y=450)

            scale7 = Scale(newSetting,from_=0 , to=700,length=200,orient=HORIZONTAL,troughcolor='#cbcf17',fg='#cf1729')
            S7 = sheet['B4']
            scale7.set(S7.value)
            scale7.place(x=355,y=520)

            scale8 = Scale(newSetting,from_=0 , to=700,length=200,orient=HORIZONTAL,troughcolor='#cbcf17',fg='#cf1729')
            S8 = sheet['C4']
            scale8.set(S8.value)
            scale8.place(x=570,y=520)
            

            global status
            status = 0 
            def saveSetting():
                global status
                status = 1  
                print(status)
                worksheet1.write('B2',scale1.get())
                worksheet1.write('C2',scale2.get())
                worksheet1.write('D2',scale3.get())
                worksheet1.write('B3',scale4.get())
                worksheet1.write('C3',scale5.get())
                worksheet1.write('D3',scale6.get())
                worksheet1.write('B4',scale7.get())
                worksheet1.write('C4',scale8.get())

                workbook1.close()

            save = Button(newSetting,text='SAVE',command=saveSetting )
            save.place(x=960,y=500)

            def stop():  
                global status
                if status == 0:
                    anwer = messagebox.askyesno("Setting not save  !","Save settings ?")
                    if anwer== True:
                        stop_vid()
                        saveSetting()
                        newSetting.destroy()
                    if anwer== False:
                        stop_vid()
                        newSetting.destroy()

                if status == 1 :
                    newSetting.destroy()
                    status = 0 

            newSetting.configure(bg='black')
            newSetting.title("Setting")
            newSetting.geometry("1056x605")
            newSetting.protocol("WM_DELETE_WINDOW",stop)

#----------------Open Image -------------------------------
id = 0
def Open():
    print("open Image ")
    openImage= Toplevel(mainWindow)
    fra = Frame(openImage, bg= 'white',width=640,height=480)
    fra.place(x=0, y=0)
    def img():
        global photos , image
        image = Image.open( "mega1\\Resur\\Image\\id_{}_".format(id)+"opencv_frame_{}.png".format(id))
        
        photos = ImageTk.PhotoImage(image)
    def show():
        global id
        id+=1
        if FileNotFoundError:
            tex_label = Label(fra,text='[Errno 2] No such file or directory')
            tex_label.grid(column=0,row=1)
            print('Erron ')
        img()
        varun_label = Label(fra,image=photos)
        varun_label.grid(column=0,row=0)
        tex_label = Label(fra,text="                             "+str(id)+"                              ")
        tex_label.grid(column=0,row=1)
        
    def back():
        global id , image
        id-=1
        if FileNotFoundError:
            tex_label = Label(fra,text='[Errno 2] No such file or directory')
            tex_label.grid(column=0,row=1)
            print('Erron ')

            if id < 0 :
                id = 0 
        img()
        varun_label = Label(fra,image=photos)
        varun_label.grid(column=0,row=0)
        tex_label = Label(fra,text="                             "+str(id)+"                              ")
        tex_label.grid(column=0,row=1)
        
    
    bt = Button(openImage,text='NEXT',command=show,width=10,height=5)
    bt.place(x=300,y=490)
    bt = Button(openImage,text='BACK',command=back,width=10,height=5)
    bt.place(x=200,y=490)

    def stop():
        openImage.destroy()
    
   
    openImage.configure(bg='black')
    openImage.title("Open Image")
    openImage.geometry("1056x605")
    openImage.protocol("WM_DELETE_WINDOW",stop)




#--------------------- draw and cap -----------------------
def getContours(imgDil1,imgContour):
 
    global globalAREA
    contours,hierarchy = cv2.findContours(imgDil1,cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_NONE)
    minarea = 1000
  
    for cnt in contours:
        for i in range(len(contours)):
            if hierarchy[0, i, 3] == -1:
                areas = cv2.contourArea(contours[i])
                if  area > minarea :
                    cv2.drawContours(imgContour, cnt, -1, (255, 0, 225), 3)
                    peri = cv2.arcLength(cnt,True)
                    #print(peri)
                    approx = cv2.approxPolyDP(cnt,0.02*peri,True)
                    x, y, w, h = cv2.boundingRect(approx)
                    #cv2.rectangle(imgContour,(x,y),(x+w,y+h),(0,255,0),2)
                    cv2.putText(imgContour,"ID:"+str(int(img_counter))+'='+str(areas),(x+5,y+15),cv2.FONT_HERSHEY_COMPLEX,0.6,(57,255,20),2)  
                    globalAREA = areas
                
    #--------------------line------------------------
books = openpyxl.load_workbook('mega1\\Setting\\SettingCap.xlsx')
SettingCap = books.active
scalelin = Scale(mainWindow,from_=0 , to=240,length=200,orient=HORIZONTAL,troughcolor='#d3dfe1',fg='#cf1729')
lin1 = SettingCap['A2']
scalelin.set(lin1.value)
scalelin.place(x=625,y=670)

scalelin2 = Scale(mainWindow,from_=0 , to=480,length=200,orient=HORIZONTAL,troughcolor='#d3dfe1',fg='#26250d')
lin2 = SettingCap['B2']
scalelin2.set(lin2.value)
scalelin2.place(x=845,y=670)       
#
scalelin4 = Scale(mainWindow,from_=0 , to=340,length=200,orient=HORIZONTAL,troughcolor='#236dc9',fg='#26250d')
lin4 = SettingCap['A3']
scalelin4.set(lin4.value)
scalelin4.place(x=625,y=740)

scalelin5 = Scale(mainWindow,from_=340, to=680,length=200,orient=HORIZONTAL,troughcolor='#236dc9',fg='#26250d')
lin5 = SettingCap['B3']
scalelin5.set(lin5.value)
scalelin5.place(x=845,y=740)

def line(imgvid,imgContour1):
   global area 

   contours,hierarchy = cv2.findContours(imgvid,cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_NONE)
   lin = scalelin.get()
   lin2 = scalelin2.get()
   lin3 = scalelin4.get()
   lin4 = scalelin5.get()
   cv2.line(imgContour1, (scalelin4.get(), scalelin.get()), (scalelin5.get(),scalelin.get()), (50, 255, 50), 2)
    
   cv2.line(imgContour1, (scalelin4.get(), scalelin2.get()), (scalelin5.get(), scalelin2.get()), (50, 255, 50), 2)


   ###########################
   minarea = 4000
   maxarea = 50000
   ##########################
   cxx = np.zeros(len(contours))
   cyy = np.zeros(len(contours))
   
   for cnt in contours:
     
        arear = cv2.contourArea(cnt)
       # print("Area =",arear)
        for i in range(len(contours)):
            if hierarchy[0, i, 3] == -1:
                area = cv2.contourArea(contours[i])
                if minarea < area < maxarea:
                    cnt = contours[i]
                    M = cv2.moments(cnt)
                    cx = int(M['m10'] / M['m00'])
                    cy = int(M['m01'] / M['m00'])
                    cv2.drawContours(imgContour1, cnt, -1, (42,167,119), 3)
                    cv2.drawMarker(imgContour1, (cx, cy), (0, 0, 255), cv2.MARKER_STAR, markerSize=10, thickness=2,line_type=cv2.LINE_AA)
                    x, y, w, h = cv2.boundingRect(cnt)
                    cv2.rectangle(imgContour1, (x, y), (x + w, y + h), (255, 0, 0), 2)
                    ID1 = img_counter
                    cv2.putText(imgContour1,"ID:"+str(ID1),(x+50,x+102),cv2.FONT_HERSHEY_COMPLEX,0.7,(0,255,0),2)
                    
                    if  lin2 > cy > lin :
                        global count
                        x, y, w, h = cv2.boundingRect(cnt)
                        cv2.rectangle(imgContour1, (x, y), (x + w, y + h), (255, 0, 0), 2)
                        cv2.putText(imgContour1, str(cx) + "," + str(cy), (cx + 10, cy + 10), cv2.FONT_HERSHEY_SIMPLEX,.3, (0, 0, 255), 1)
                        cv2.putText(imgContour1,"ID:"+str(img_counter),(cx-10, cy-30 ),cv2.FONT_HERSHEY_COMPLEX,0.7,(0,255,0),2)
                        cv2.drawMarker(imgContour1, (cx, cy), (0, 0, 255), cv2.MARKER_STAR, markerSize=5, thickness=1,line_type=cv2.LINE_AA)
                        cv2.line(imgContour1, (scalelin4.get(), lin), (scalelin5.get(), lin), (255, 0, 0), 5)
                        cv2.line(imgContour1, (scalelin4.get(), lin2), (scalelin5.get(), lin2), (255, 0, 0), 5)
                        count =1  

                       
                    #if cy > lineypos2  :
                       # x, y, w, h = cv2.boundingRect(cnt)
                        #cv2.rectangle(imgContour1, (x, y), (x + w, y + h), (255, 0, 0), 2)
                        #cv2.putText(imgContour1,"ID:"+str(img_counter)+"  Area:"+str(int(area)),(x+50,x+102),cv2.FONT_HERSHEY_COMPLEX,0.7,(0,255,0),2)   
                        #cv2.line(imgContour1, (0, lineypos2), (width, lineypos2), (255, 0, 255), 5)



#-------------------Block and Save  & Unblock -----------------------------------
def blocks():
    global blc
    workbook1s = xlsxwriter.Workbook('mega1\\Setting\\SettingCap.xlsx')
    worksheet1 = workbook1s.add_worksheet("SettingCap")
    worksheet1.write('A2',scalelin.get())
    worksheet1.write('B2',scalelin2.get())
    worksheet1.write('A3',scalelin4.get())
    worksheet1.write('B3',scalelin5.get())
    workbook1s.close()
    print(scalelin.get(),scalelin2.get(),scalelin4.get(),scalelin5.get())
    print('blcok scale')
    blc = 'disable'
    scalelin.config(state=blc,takefocus=0)
    scalelin2.config(state=blc,takefocus=0)
    scalelin4.config(state=blc,takefocus=0)
    scalelin5.config(state=blc,takefocus=0)
def unblock():
    global blc
    print('UnBlock scale')
    ublc = 'active'
    scalelin.config(state=ublc,takefocus=0)
    scalelin2.config(state=ublc,takefocus=0)
    scalelin4.config(state=ublc,takefocus=0)
    scalelin5.config(state=ublc,takefocus=0)
bloc = Button(mainWindow, text= 'BLOCK & SAVE', command=blocks)
bloc.place(x=1080,y=715)
Unloc = Button(mainWindow, text= 'UNBLOCK',command=unblock)
Unloc.place(x=1180, y=715)



#-------------cap--------------------
def caps():
    global count 
    global wri
    global img_counter 
    if count == 1 :
        img_counter = img_counter+1 
        img_name = "mega1\\Resur\\Image\\id_{}_".format(img_counter)+"opencv_frame_{}.png".format(img_counter)
        cv2.imwrite(img_name,res)
        print("{} written!".format(img_name))
        showResur()    
        wri +=1
        couter()
        write()
        time.sleep(0.5) 
        count = 0
    uncouter()
        
        
        
       
        

#----------------Excel data --------------------------------------
tex = Frame(mainWindow,bg= 'black', height =0, width =0 )
tex.place(x=510,y=570) 


def write():
    global t 
    t = datetime.datetime.now()
    weigh()
   
    worksheet.write('A1',"ID " )
    worksheet.write('B1',"Area" )
    worksheet.write('C1',"Weight")
    worksheet.write('D1',"Date-Time")
    worksheet.write('A{}'.format(wri),"Tôm -id_{}:".format(img_counter) )
    worksheet.write('B{}'.format(wri),globalAREA )
    worksheet.write('C{}'.format(wri),y )
    worksheet.write('D{}'.format(wri),t.strftime('%X %x') )

    if WeightStatus ==1 :
        wights = Entry(mainWindow,font=("Arial",14))
        wights.place(x=625,y=630)
        texwight=Label(mainWindow,text="ID :"+str((wri-1)),font="Arial")
        texwight.place(x=500,y=630)
        worksheet.write('G1', "Real Weight")
        def wr():
            if wights.get() == "non":
                worksheet.write('G{}'.format(wri), "No data")
            else:
                worksheet.write('G{}'.format(wri), float(wights.get()))
            wights.delete(0, 'end')
        def wrenter(e):
            if wights.get() == "non":
                worksheet.write('G{}'.format(wri), "No data")
            else:
                worksheet.write('G{}'.format(wri), float(wights.get()))
            wights.delete(0, 'end')

        bton = Button(text= "Add Wights",command=wr)
        bton.place(x=860,y=630)

        mainWindow.bind('<Return>',wrenter)
        

    lbTex = Label(tex,text="Tôm ID-{}   Area = ".format(img_counter)+str(globalAREA)
                        +'   weigh = {}'.format(y)+"         " ,font=('Comic Sans',15,'bold'), 
                                fg='#42ecf5',bg='white',bd=10,
                        padx=10)         
    lbTex.grid(row=0, column=0)
    lbTex = Label(tex,text='Số Lượng :'+str(img_counter),font=('Comic Sans',15,'bold'),fg='#42ecf5',bg='white',
                        bd=10,
                        padx=10)
    lbTex.grid(row=0, column=1)



def formulas():
    if messagebox.showwarning(title="Warnig",message="Changing the value can affect the accuracy of the program"):
        print('Add formual ')
        formula= Toplevel(mainWindow)
        
        load = Image.open("mega1\\Setting\\Formula\\formula.jpg")
        render = ImageTk.PhotoImage(load)
        img = Label(formula, image=render)
        img.image = render
        img.place(x=90, y=0)
        
        def submit():
            workbook2 = xlsxwriter.Workbook('mega1\\Setting\\Formula\\Formula.xlsx')
            worksheet2 = workbook2.add_worksheet("SettingCap")
            worksheet2.write('A1',float(A.get()))
            worksheet2.write('B1',float(B.get()))
            worksheet2.write('C1',float(C.get()))
            worksheet2.write('D1',float(D.get()))
            worksheet2.write('E1',float(E.get()))
            worksheet2.write('F1',float(F.get()))
            workbook2.close()
            labelF()
        def changeScw(e):
            workbook2 = xlsxwriter.Workbook('mega1\\Setting\\Formula\\scale.xlsx')
            worksheet2 = workbook2.add_worksheet("sheet1")
            worksheet2.write('G1',float(ScaleW.get()))
            workbook2.close()
            labelF()

        A = Entry(formula,font=("Arial",15))
        A.place(x=60,y=130)
        texA=Label(formula,text="A",font="Arial")
        texA.place(x=30,y=130)

        B = Entry(formula,font=("Arial",15))
        B.place(x=60,y=160)
        texB=Label(formula,text="B",font="Arial")
        texB.place(x=30,y=160)

        C = Entry(formula,font=("Arial",15))
        C.place(x=60,y=190)
        texC=Label(formula,text="C",font="Arial")
        texC.place(x=30,y=190)

        D = Entry(formula,font=("Arial",15))
        D.place(x=60,y=220)
        texD=Label(formula,text="D",font="Arial")
        texD.place(x=30,y=220)

        E = Entry(formula,font=("Arial",15))
        E.place(x=60,y=250)
        texE=Label(formula,text="E",font="Arial")
        texE.place(x=30,y=250)

        F = Entry(formula,font=("Arial",15))
        F.place(x=60,y=280)
        texF=Label(formula,text="F",font="Arial")
        texF.place(x=30,y=280)

        ScaleW = Entry(formula,font=("Arial",15))
        ScaleW.place(x=300,y=410)


        def labelF():
            weigh()
            
            texformulal=Label(formula,text="A = "+str(As)+" | B = "+str(Bs)+" | C = "+str(Cs)+" | D = "+str(Ds)+" | E = "+str(Es)+" | F = "+str(Fs)+"|                        ",font="Arial")
            texformulal.place(x=30,y=330)
            texF=Label(formula,text="Scale = "+str(Scw)+"    ",font="Arial")
            texF.place(x=100,y=410)
        submit_button=Button(formula,text="Change",command=submit)
        submit_button.place(x=300,y=200)

        formula.bind('<Return>',changeScw)
        labelF()
        def weight():
            global WeightStatus
            WeightStatus = 1 
            print(WeightStatus)

        submit_button=Button(formula,text=" Mode |Enter weight data|",command=weight)
        submit_button.place(x=0,y=380)


        
        def stop():
            formula.destroy()
        formula.configure(bg='black')
        formula.title("Add Formula")
        formula.geometry("640x480")
        formula.protocol("WM_DELETE_WINDOW",stop)

def weigh():
    global y ,As,Bs,Cs,Ds,Es,Fs ,Scw
    bookx = openpyxl.load_workbook('mega1\\Setting\\Formula\\Formula.xlsx')
    formulx= bookx.active

    bookscalex = openpyxl.load_workbook('mega1\\Setting\\Formula\\scale.xlsx')
    scalex=bookscalex.active
    Scw = scalex['G1']
    Scw=Scw.value

    x =  globalAREA/Scw
    print(Scw)
    A = formulx['A1']
    As=A.value

    B = formulx['B1']
    Bs=B.value

    C = formulx['C1']
    Cs=C.value

    D = formulx['D1']
    Ds=D.value

    E = formulx['E1']
    Es=E.value

    F = formulx['F1']
    Fs=F.value



    print(As,Bs,Cs,Ds,Es,Fs)
    y = (As)*pow(x,5)+(Bs)*pow(x,4)+(Cs)*pow(x,3)+(Ds)*pow(x,2)+(Es)*pow(x,1)+(Fs)    
    y = round(y,2)
#-----------------Setup Camera -------------
def setupcam():
    if active =='disable':
        if messagebox.showerror(title="Erron",message="Camera ID not selected"):
            print(idcam)
    else:
        setup= Toplevel(mainWindow)
        frame = Frame(setup, bg= 'black',width=1000,height=800)
        frame.place(x=0, y=0)
        F_lbl = Label(frame)
        F_lbl.grid(row=0, column=0)
        stop_vid()
        def show():
            if cam_on:
                ret, imgvid = cap.read()
                imgvid = imgvid[scales.get():scale3s.get(),scale2s.get():scale4s.get()]
                camer =cv2.cvtColor(imgvid, cv2.COLOR_BGR2RGB)
                imgBlur = cv2.GaussianBlur(camer,(7,7),1)
                imgv = Image.fromarray(camer)
                imgtk = ImageTk.PhotoImage(image=imgv) 
                F_lbl.imgtk = imgtk    
                F_lbl.configure(image=imgtk) 
                setup.after(20,show)
    def start():
        global cam_on, cap
        cam_on = True
        cap = cv2.VideoCapture(idcam)
        show() 

    def save():
            workbook2 = xlsxwriter.Workbook('mega1\\Setting\\SttingCap.xlsx')
            worksheet2 = workbook2.add_worksheet("SettingCap")
            worksheet2.write('A5',int(scales.get()))
            worksheet2.write('B5',int(scale2s.get()))
            worksheet2.write('A6',int(scale3s.get()))
            worksheet2.write('B6',int(scale4s.get()))
            workbook2.close()
   
 
    starts= Button(setup,text='Start Camera',command=start)
    starts.place(x=0,y=600)
    saves = Button(setup,text='save',command=save,width=10)
    saves.place(x=100,y=600)
    

    value()
    scales = Scale(setup,from_=0 , to=480,length=200,orient=HORIZONTAL,troughcolor='#cbcf17',fg='#cf1729')
    scales.set(y1)
    scales.place(x=30,y=650)
    
    scale2s = Scale(setup,from_=0 , to=680,length=200,orient=HORIZONTAL,troughcolor='#cbcf17',fg='#cf1729')
    scale2s.set(x1)
    scale2s.place(x=30,y=720)

    scale3s = Scale(setup,from_=480 , to=0,length=200,orient=HORIZONTAL,troughcolor='#cbcf17',fg='#cf1729')

    scale3s.set(y2)
    scale3s.place(x=250,y=650)
    scale4s= Scale(setup,from_=680 , to=0,length=200,orient=HORIZONTAL,troughcolor='#cbcf17',fg='#cf1729')

    scale4s.set(x2)
    scale4s.place(x=250,y=720)

    def stop():
        stop_vid()
        setup.destroy()
    setup.configure(bg='black')
    setup.title("Setup Camera")
    setup.geometry("1000x800")
    setup.protocol("WM_DELETE_WINDOW",stop)
    

def value():
        global x1 ,x2 ,y1,y2
        book = openpyxl.load_workbook('mega1\\Setting\\SttingCap.xlsx')
        sheet = book.active

        y1 = sheet['A5']
        y1=y1.value
        x1 = sheet['B5']
        x1=x1.value
        y2 = sheet['A6']
        y2=(y2.value)
        x2 = sheet['B6']
        x2=x2.value



#--------------------- QUIT MAIN     --------- 
def quits():
    print('Quit main , see you again !!!!!!!!!')
    if messagebox.askokcancel("Save and Close", "Do you want to quit?"):
        workbook.close()
        mainWindow.destroy()
##############################################
menubar =Menu(mainWindow)
mainWindow.config(menu=menubar)

toolMenu  =Menu(menubar,tearoff=0,activebackground='#486986',bg='#e2aeb1')
menubar.add_cascade(label="Option ",menu=toolMenu)
toolMenu.add_command(label="Setup Camera",command=setupcam)
toolMenu.add_command(label="Setting",command=setting)
toolMenu.add_command(label="Oppen Image",command=Open)
toolMenu.add_command(label="Add Formula",command=formulas)
toolMenu.add_separator()
toolMenu.add_command(label="Exit",command= quits)
##############################################

mainWindow.protocol("WM_DELETE_WINDOW",quits)
mainWindow.mainloop()
